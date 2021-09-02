import openpyxl
from openpyxl.styles import PatternFill
from colorama import init 
from termcolor import colored
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import Font

wb=openpyxl.load_workbook('ExportList_esxi_dns.xlsx') #File1 with checking column
wb2=openpyxl.load_workbook('Audit VMware DNS Prod v1.0.xlsx') #File2 with we compare
sh=wb['ExportList_esxi_dns'] 
sh2=wb2['GlobalSite']

def for_ok():
    for i in range(2,sh.max_row+1):
        a=sh.cell(row=i,column=1).value
        for j in range(2,sh2.max_row+1):
            b=sh2.cell(row=j,column=3).value
            if a == b:
                print(colored(a, 'green', 'on_red')) 
                font = Font(color="00FF00")
                sh.cell(row=i,column=3).value='ok'
                sh.cell(row=i,column=3).font = Font(color="096A09")
                sh.cell(row=i,column=3).fill=PatternFill(start_color="B0F2B6", end_color="B0F2B6", fill_type = "solid")
                #sh.cell(row=i,column=3).fill=PatternFill(start_color="82C46C", end_color="82C46C", fill_type = "solid")


def for_nok():
    for k in range(2,sh.max_row+1):
        if sh.cell(row=k,column=3).value is None:
            sh.cell(row=k,column=3).value='nok'
            sh.cell(row=k,column=3).font=Font(color="FF0921")
            sh.cell(row=k,column=3).fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")

for_ok()
for_nok()


wb.save('check_done.xlsx')