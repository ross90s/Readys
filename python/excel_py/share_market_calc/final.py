import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

wb=openpyxl.load_workbook('Account_23-04-2020.xlsx')
wb2=openpyxl.load_workbook('op_account.xlsx')
sh=wb['Compte']
sh2=wb2['Feuil1']

values= []    # Empty list to store non duplicate names of companies

for i in range(1,sh.max_row+1):     
    a=sh.cell(row=i,column=4).value
    if a is None:
        pass
    else:
        if a in values:  # If name is already in 'values' do nothing
            pass
        else:
            values.append(sh.cell(row=i,column=4).value)   # Else add the name in 'values'

for x in range(len(values)):         #Loop to add elements of 'values' in column1 of op_account.xlsx
    if values[x] is None:
        pass
    else:                
        sh2.cell(row=x+1,column=1).value=values[x]

wb2.save('op_account.xlsx')     # save the  output file

#Loop to print total invested amount in front of company
for i in range(2,sh2.max_row+1):  # Insert company one by one from column1 of op_account.xlsx
    invest=0
    for j in range(2,sh.max_row+1): 
        c=sh2.cell(row=i,column=1).value
        if(c==sh.cell(row=j,column=4).value):
            if sh.cell(row=j,column=9).value is None:
                sh.cell(row=j,column=9).value=0
            else:
                invest=invest+float(sh.cell(row=j,column=9).value)
    if sh2.cell(row=i,column=1).value is None:
        pass
    else:                
        sh2.cell(row=i,column=2).value=invest
        print("Total invested in",c,"=",invest,end=' \n')
    
wb2.save('op_account.xlsx')     # save the  output file

########## Insert table of extra charges  ###############

charge_list=['Frais de courtage','Taxe sur les Transactions Financières (TTF)','Variation Fonds Monétaires (EUR)','Frais de connexion aux places boursières 2020 (New York Stock Exchange - NSY)','Frais de connexion aux places boursières 2020 (NASDAQ - NDQ)','Frais de connexion aux places boursières 2020 (Euronext Amsterdam - EAM)','Frais de connexion aux places boursières 2020 (Xetra - XET)']

for x in range(len(charge_list)):             #Loop to add elements of 'charge_list' in column5 of op_account.xlsx
    sh2.cell(row=x+2,column=5).value=charge_list[x]

wb2.save('op_account.xlsx')     # save the  output file

print()
print('###  Different CHARGES   ###')
for i in range(2,len(charge_list)+2):  # Insert charge_type one by one from column5 of op_account.xlsx
    charges=0
    for j in range(2,sh.max_row+1): # Compare string in column5 of op_account.xlsx to column6 of Account.xlsx  
        ch=sh2.cell(row=i,column=5).value
        if(ch==sh.cell(row=j,column=6).value):
            if sh.cell(row=j,column=9).value is None:
                sh.cell(row=j,column=9).value=0
            else:
                charges=charges+float(sh.cell(row=j,column=9).value)
    print("Total charges of ",ch,"=",charges,end=' \n')
    sh2.cell(row=i,column=6).value=charges
all_ch=0
for k in range(2,len(charge_list)+1):
    all_ch=all_ch+float(sh2.cell(row=k,column=6).value)
print('Total of charge =',all_ch)
sh2['F9'] = '=SUM(F2:F8)'
sh2['E9'] = "Total of charges"

####### Total In/Out balance ###########

funds=['Versement de fonds','Retrait de fonds']

for m in range(len(funds)):             #Loop to add elements of 'Transfers' in column5 of op_account.xlsx
    sh2.cell(row=m+17,column=5).value=funds[m]

wb2.save('op_account.xlsx')     # save the  output file

print()
print('###  Total In/Out   ###')
print()
for i in range(17,19):  # compare transfer_type one by one from column5 of op_account.xlsx
    tot=0
    for j in range(2,sh.max_row+1):              # Compare string in column5 of op_account.xlsx to column6 of Account.xlsx  
        ch=sh2.cell(row=i,column=5).value      # ch=1st string in funds ie row 10 column 5  
        if(ch==sh.cell(row=j,column=6).value):
            if sh.cell(row=j,column=9).value is None:
                sh.cell(row=j,column=9).value=0
            else:
                tot=tot+float(sh.cell(row=j,column=9).value)
    print("Total of ",ch,"=",tot,end=' \n')
    sh2.cell(row=i,column=6).value=tot

all_funds=0
for k in range(17,19):
    if sh2.cell(row=k,column=6).value is None:
        sh2.cell(row=k,column=6).value=0
    else:
        all_funds=all_funds+float(sh2.cell(row=k,column=6).value)
print('Total investment =',all_funds)

sh2['E19'] = "Total sum invested"
sh2['F19'] = '=SUM(F17:F18)'

# Add headers in A and B columns
data=["Company Name","Investment"]
for i in range(len(data)):
    sh2.cell(row=1,column=i+1).value=data[i]

# Add headers in E and F columns
data1=["Charge Type","Amount"]
for i in range(len(data1)):
    sh2.cell(row=1,column=i+5).value=data1[i]

# Add headers in E16 and F16 ie below above table
data2=["Transfer Type","Amount"]
for i in range(len(data2)):
    sh2.cell(row=16,column=i+5).value=data2[i]

# To make font bold for selected cells
bold_list=['A1','B1','E1','F1','E9','F9','E16','F16','E19','F19']
for b in bold_list:
    sh2[b].font = Font(bold=True)

#for b in bold_list:
#    sh2[b].font = GLOBAL_TITLE_FONT

#GLOBAL_TITLE_FONT = Font(name='Calibri',
#                         size=24,
#                         bold=True,
#                         color='e64117')

bg_list=['A1','B1','E1','F1','E16','F16']
for b in bg_list:
    sh2[b].fill= PatternFill(start_color="9999FF",
                             end_color="9999FF",
                             fill_type="solid")

wb2.save('op_account.xlsx')     # save the  output file

