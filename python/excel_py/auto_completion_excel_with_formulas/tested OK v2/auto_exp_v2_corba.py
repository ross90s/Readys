import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import datetime

wb = openpyxl.load_workbook('v2_old.xlsx')
#wb = openpyxl.load_workbook('test3.xlsx', data_only = 'True')
wb2=openpyxl.load_workbook('v2_update_file.xlsx')
#wb = openpyxl.load_workbook('Trappes.xlsx', data_only=True)
#wb = openpyxl.load_workbook('Trappes.xlsx', read_only=True, data_only=True)        
#wb = openpyxl.load_workbook('Trappes.xlsx', data_only = True)
#wb=openpyxl.load_workbook('Trappes.xlsx')
sh=wb['Readme']
sh2=wb2['Feuil1']
sh3=wb['VNF-VIM']
sh4=wb['NIC']
sh5=wb['Network']
sh6=wb['StorageVolume']
sh7=wb['Net-VDS']
sh8=wb['ComputePool']
sh9=wb['Compute']
sh10=wb['VMwareCluster']
sh11=wb['User']
sh12=wb['DRSrule']

####### ReadMe Sheet

#Write current date in 1st column 
def date_update():
    for i in range(68,sh.max_row+1):   
        a=sh.cell(row=i,column=1).value
        #print(a)
        if a is None:
            sh.cell(row=i,column=1).value=datetime.datetime.now()
            break

#Write version
def version_maj():
    for i in range(68,sh.max_row+1):
        a=sh.cell(row=i,column=2).value
        #b=float(sh.cell(row=i-1,column=2).value)
        if a is None:
            b=float(sh.cell(row=i-1,column=2).value)
            print(b)
            sh.cell(row=i,column=2).value=(b+0.01)
            break

#Modif details
def modif_detail():
    for i in range(68,sh.max_row+1):   
        a=sh.cell(row=i,column=3).value
        #print(a)
        if a is None:
            sh.cell(row=i,column=3).value=sh2.cell(row=2,column=1).value
            break           

#Modif Trigramme
def trigramme():
    for i in range(68,sh.max_row+1):   
        a=sh.cell(row=i,column=8).value
        if a is None:
            #print(sh2.cell(row=2,column=2).value)
            sh.cell(row=i,column=8).value=sh2.cell(row=2,column=2).value
            break 


##### VNF-VIM Sheet

#Site-name
def site_name():
    for i in range(5,sh3.max_row+1): 
        a=sh3.cell(row=i,column=2).value 
        b=sh3.cell(row=i-1,column=2).value
        if a is None:
            sh3.cell(row=i,column=2).value=b
            break  

#True_false
def true_false():
    for i in range(5,sh3.max_row+1): 
        a=sh3.cell(row=i,column=3).value 
        b=sh3.cell(row=i-1,column=3).value
        if a is None:
            sh3.cell(row=i,column=3).value=b
            break  

#vnf name: First 3-6 rows are None, so start count from 8
def vnf_name():
    for i in range(8,sh3.max_row+1):
        a=sh3.cell(row=i,column=4).value
        if a is None:
            sh3.cell(row=i,column=4).value=sh2.cell(row=2,column=3).value
            break

#Virtual type:First 3-6 rows are None, so start count from 8
def virt_type():
    for i in range(8,sh3.max_row+1): 
        a=sh3.cell(row=i,column=6).value 
        b=sh3.cell(row=i-1,column=6).value # If VMware65 then ok, otherwise check after run
        if a is None:
            sh3.cell(row=i,column=6).value=b
            break 

#id_int
def id_int():
    for i in range(8,sh3.max_row+1):
        a=sh3.cell(row=i,column=7).value
        if a is None:
            print('1')
            sh3.cell(row=i,column=7).value=sh2.cell(row=2,column=4).value
            break  


##### NIC sheet



#Count Datastore
count=0
for x in sh2['J']:
    if x.value == 'datastore-fcoe':
        count=count+1
#print('Count',count)


def vnf_nic():
    for i in range(3,sh4.max_row+1):
        a=sh4.cell(row=i,column=2).value
        if a is None:
            print(i)
            for j in sh3['D']:
                if j.value ==sh2.cell(row=2,column=3).value:
                    print(j.coordinate)
                    stri=j.coordinate
                    stri=stri.replace("D","A")
                    print(stri)
                    z="='VNF-VIM'!{d}".format(d=stri)
                    print(z)
                    if sh2.cell(row=2,column=5).value==2:
                        sh4.cell(row=i,column=2).value="='VNF-VIM'!{d}".format(d=stri)
                        sh4.cell(row=i+1,column=2).value="='VNF-VIM'!{d}".format(d=stri)
                        print(2)
                    else:
                        sh4.cell(row=i,column=2).value="='VNF-VIM'!{d}".format(d=stri)
                        print(1)
                    break
            break

def new_vnf_nic():
    
    for i in range(3,sh4.max_row+1):
        a=sh4.cell(row=i,column=2).value
        if a is None:
            if sh2.cell(row=2,column=5).value==2:
                sh4.cell(row=i,column=2).value="='VNF-VIM'!{d}".format(d=stri)
                sh4.cell(row=i+1,column=2).value="='VNF-VIM'!{d}".format(d=stri)
                print(2)
            else:
                sh4.cell(row=i,column=2).value="='VNF-VIM'!{d}".format(d=stri)
                print(1)
            break


#NIC-AdapterPolicy
def nic_ada():
    for i in range(2,sh4.max_row+1):
        a=sh4.cell(row=i,column=3).value
        b=sh4.cell(row=i-1,column=3).value
        if a is None:
            if sh2.cell(row=2,column=5).value==2:
                print("nic is 2")
                sh4.cell(row=i,column=3).value=b
                print('nic_ada',b)
                sh4.cell(row=i+1,column=3).value=b
            else:
                sh4.cell(row=i,column=3).value=b
            break       

#NIC-compute_NICs
def nic_comp():
    for i in range(2,sh4.max_row+1):
        a=sh4.cell(row=i,column=4).value
        if a is None:
            if sh2.cell(row=2,column=5).value==2:
                sh4.cell(row=i,column=4).value='vnic5'
                sh4.cell(row=i+1,column=4).value='vnic6'
            else:
                sh4.cell(row=i,column=4).value='vnic5'
            break


##### Network sheet

def net_name_id():
    for i in range(2,count1+1):
        a=sh2.cell(row=i,column=6).value.split()[1][:-1]
        b=sh2.cell(row=i,column=6).value.split()[2]
        subnet=sh2.cell(row=i,column=7).value
        for j in range(22,sh5.max_row+1):
            c=sh5.cell(row=j,column=3).value
            if c is None:
                sh5.cell(row=j,column=3).value=a
                sh5.cell(row=j,column=5).value=b
                sh5.cell(row=j,column=2).value="='VNF-VIM'!{d}".format(d=stri)
                sh5.cell(row=j,column=7).value='vds'
                sh5.cell(row=j,column=9).value=subnet
                if 'NFS' not in a:
                    sh5.cell(row=j,column=6).value=1500
                    sh5.cell(row=j,column=8).value="=NIC!{d}".format(d=stri1)
                else:
                    sh5.cell(row=j,column=6).value=9000  
                    sh5.cell(row=j,column=11).value=sh2.cell(row=i,column=7).value[:-4]+'220'  
                    sh5.cell(row=j,column=12).value=sh2.cell(row=i,column=7).value[:-4]+'240'
                    sh5.cell(row=j,column=8).value="=NIC!{d}".format(d=stri2)
                if a is None:
                    break
                break



### StorageVolume sheet

def descr_char():
    for i in range(2,sh2.max_row): 
        a=sh2.cell(row=i,column=8).value
        b=sh2.cell(row=i,column=9).value
        c=sh2.cell(row=i,column=10).value
        d=sh2.cell(row=i,column=11).value
        for j in range(2,sh6.max_row):
            e=sh6.cell(row=j,column=3).value
            if e is None:               
                sh6.cell(row=j,column=3).value=a
                sh6.cell(row=j,column=4).value=b
                sh6.cell(row=j,column=5).value=c
                sh6.cell(row=j,column=6).value=d
                
                if sh6.cell(row=j,column=5).value == 'datastore-fcoe':
                    f=sh2.cell(row=2,column=3).value.lower()+'-'+'ds'+'0'+str(i-1)
                    sh6.cell(row=j,column=7).value=f
                if sh6.cell(row=j,column=5).value == 'nas-nfs':
                    for v in sh5['C']:
                        if v.value is not None:
                            if 'NFS' in v.value and sh2.cell(row=2,column=3).value.upper() in v.value :
                                c=v.coordinate
                                c=c.replace("C","A")
                                print("c=of datastore",c)
                                sh6.cell(row=j,column=8).value="='Network'!{d}".format(d=c)
                #print('j',j)
                if sh6.cell(row=j,column=6).value is not None:
                    sh6.cell(row=j,column=2).value=z
                if a is None:
                    break
                break

### Net-VDS

def net_vds():
    for j in sh5['C']:
        if j.value is not None:
            if sh2.cell(row=2,column=3).value.lower() in j.value.lower():
                print(j.coordinate)
                for m in range(2,sh7.max_row):
                    if sh7.cell(row=m,column=2).value is None:
                        c=j.coordinate
                        c=c.replace("C","A")
                        print("c=",c)
                        sh7.cell(row=m,column=2).value="='Network'!{d}".format(d=c)
                        if 'NFS' in j.value:
                            sh7.cell(row=m,column=3).value=sh2.cell(row=2,column=3).value.lower()+'-vsw02'
                        else:
                            sh7.cell(row=m,column=3).value=sh2.cell(row=2,column=3).value.lower()+'-vsw01'
                        pg_string=j.value.replace(sh2.cell(row=2,column=3).value.upper(),"")
                        pg_last=pg_string[:-1]
                        sh7.cell(row=m,column=4).value='PG_'+sh2.cell(row=2,column=3).value.upper()+'_'+pg_last
                        sh7.cell(row=m,column=5).value='access'
                        break


#### Compute-Pool                         

def compute_pool():
    for k in range(2,sh8.max_row):
        if sh8.cell(row=k,column=2).value is None:
            sh8.cell(row=k,column=2).value="='VNF-VIM'!{d}".format(d=stri)
            sh8.cell(row=k,column=3).value=sh2.cell(row=2,column=3).value.lower()+'_65'
            break
        
###### Compute

def compute():
    for k1 in range(2,sh2.max_row+1):
        if sh2.cell(row=k1,column=12).value is not None:
            for k in range(2,sh9.max_row):
                if sh9.cell(row=k,column=3).value is None:
                    sh9.cell(row=k,column=3).value=sh2.cell(row=k1,column=12).value.lower()
                    sh9.cell(row=k,column=4).value=sh2.cell(row=k1,column=13).value
                    sh9.cell(row=k,column=5).value=sh2.cell(row=k1,column=14).value
                    if sh9.cell(row=k,column=3).value is not None:
                       sh9.cell(row=k,column=2).value="='ComputePool'!{d}".format(d=key_pool)    
                    break


#**********   Checking done *************
# #### VMwareCluster

def vmware_cluster():
    for t in range(2,sh10.max_row):
        if sh10.cell(row=t,column=3).value is None:
            sh10.cell(row=t,column=3).value = sh2.cell(row=2,column=3).value
            sh10.cell(row=t,column=4).value = sh10.cell(row=t-1,column=4).value
            sh10.cell(row=t,column=5).value = sh10.cell(row=t-1,column=5).value
            sh10.cell(row=t,column=7).value = sh2.cell(row=2,column=16).value
            sh10.cell(row=t,column=2).value = "='VNF-VIM'!{d}".format(d=stri)
            sh10.cell(row=t,column=8).value = "='ComputePool'!{d}".format(d=key_pool)
            if sh2.cell(row=2,column=15).value == 'true':
                sh10.cell(row=t,column=6).value = 1
            break

##### DRS Rule

def drsrule():
    for e in range(2,sh2.max_row):
        if sh2.cell(row=e,column=17).value is not None:
            for g in range(2,sh12.max_row):
                if sh12.cell(row=g,column=5).value is None:
                    sh12.cell(row=g,column=5).value = sh2.cell(row=e,column=17).value


##### users

def users():
    for i in range(1,3):
        for j in range(2,sh11.max_row):
            if sh11.cell(row=j,column=3).value is None:
                print('i','j',i,j)
                sh11.cell(row=j,column=3).value=sh2.cell(row=2,column=3).value.lower()+'-'+'0'+str(i)
                sh11.cell(row=j,column=2).value="='VNF-VIM'!{d}".format(d=stri)
                break
    for i in range(1,3):
        for j in range(2,sh11.max_row):
            if sh11.cell(row=j,column=3).value is None:
                print('i','j',i,j)
                sh11.cell(row=j,column=3).value='sfr'+'-'+sh2.cell(row=2,column=3).value.lower()+'-'+'0'+str(i)
                sh11.cell(row=j,column=2).value="='VNF-VIM'!{d}".format(d=stri)
                break
    


###### Readme           

modif_detail()
date_update()
version_maj()
trigramme()

####VNF

site_name()
true_false()
vnf_name()
virt_type()
id_int()

#vnf of NIC
for j in sh3['D']:
    if j.value == sh2.cell(row=2,column=3).value:
        print(j.coordinate)
        stri=j.coordinate
        stri=stri.replace("D","A")
        print("stri=",stri)
        z="='VNF-VIM'!{d}".format(d=stri)
        print(z)
        break


###NIC  

new_vnf_nic()
nic_ada()
nic_comp()

#To get last vnic5 & vnic6

c1=[]
for k in sh4['D']:
    if k.value == 'vnic5':
        c1.append(k.coordinate) 
        print('c1=',c1)
stri1=c1[-1]
stri1=stri1.replace("D","A")
print("str1=",stri1)

c2=[]
for k in sh4['D']:
    if k.value == 'vnic6':
        c2.append(k.coordinate) 
        #print(c1)
stri2=c2[-1]
stri2=stri2.replace("D","A")
print("str2=",stri2)

#Count Network and use in net_name_id()
count1=0
for x1 in sh2['F']:
    if x1.value is not None:
        count1=count1+1
print('count1=',count1)
####NETWORK  

net_name_id()

###### StorageVolume sheet
descr_char()

######## Net-VDS
net_vds()

###### Compute Pool
compute_pool()

#### Compute
for k2 in sh8['C']:
    if k2.value is not None:
        if sh2.cell(row=2,column=3).value.lower() in k2.value:
            key_pool=k2.coordinate
            key_pool=key_pool.replace("C","A")
            print("key_pool:",key_pool)

compute()

#**********   Checking done *************
##### VMwareCluster
vmware_cluster()

#### DRSRULE ####
#drsrule()

##users
users()

wb.save('new_version.xlsx')